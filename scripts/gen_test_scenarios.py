"""
scripts/gen_test_scenarios.py
Generate docs/medsignal_test_scenarios.xlsx from the full test catalogue.

Run: poetry run python scripts/gen_test_scenarios.py
"""

import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Colour palette ──────────────────────────────────────────────────────────
GREEN_DARK  = "1B5E20"
GREEN_MID   = "2E7D32"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def bold_font(size=11, color="FFFFFF"):
    return Font(bold=True, size=size, color=color)

def thin_border():
    side = Side(style="thin", color="BDBDBD")
    return Border(left=side, right=side, top=side, bottom=side)

def center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left_align(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)


# ── Test catalogue ───────────────────────────────────────────────────────────
# Columns: ID | Component | Test File | Test Function | Description | Path Type | Level | Marker | Notes
TESTS = [
    # ── Agent 1 ──────────────────────────────────────────────────────────────
    ("A1-01","Agent 1 – Signal Detector","unit/test_agent1.py","test_golden_signal_dupilumab",
     "GPT-4o generates 3 valid PubMed queries for dupilumab x conjunctivitis (golden signal); stat_score from state is not overwritten.",
     "Happy Path","Unit","none",""),

    ("A1-02","Agent 1 – Signal Detector","unit/test_agent1.py","test_high_severity_signal",
     "Generates valid queries for gabapentin x cardio-respiratory arrest (death_count=67, high-severity).",
     "Happy Path","Unit","none",""),

    ("A1-03","Agent 1 – Signal Detector","unit/test_agent1.py","test_borderline_prr_signal",
     "PRR just at threshold (2.0) — minimal statistical strength — still produces 3 distinct queries.",
     "Edge Case / Boundary","Unit","none","Boundary: PRR=2.0"),

    ("A1-04","Agent 1 – Signal Detector","unit/test_agent1.py","test_high_volume_signal",
     "semaglutide x nausea with large case count — queries generated correctly.",
     "Happy Path","Unit","none",""),

    ("A1-05","Agent 1 – Signal Detector","unit/test_agent1.py","test_no_outcome_flags",
     "All outcome flags (deaths, hosp, LT) are zero — queries still generated.",
     "Edge Case","Unit","none",""),

    ("A1-06","Agent 1 – Signal Detector","unit/test_agent1.py","test_template_fallback",
     "When GPT-4o is unavailable, _template_queries() returns 3 distinct fallback queries covering mechanistic, epidemiological, and clinical angles.",
     "Happy Path","Unit","none","No API call"),

    ("A1-07","Agent 1 – Signal Detector","unit/test_agent1.py","test_query_uniqueness",
     "All 3 queries returned by GPT-4o are distinct (no duplicate queries).",
     "Happy Path","Unit","none",""),

    ("A1-08","Agent 1 – Signal Detector","unit/test_agent1.py","test_all_golden_drugs",
     "All 10 golden-signal drugs produce 3 valid queries using template fallback — no crash for any drug.",
     "Happy Path","Unit","none","Template; no API cost"),

    # ── Agent 2 ──────────────────────────────────────────────────────────────
    ("A2-01","Agent 2 – Literature Retriever","unit/test_agent2.py","test_lit_score_empty",
     "Empty abstracts list returns LitScore = 0.0 exactly.",
     "Edge Case","Unit","unit",""),

    ("A2-02","Agent 2 – Literature Retriever","unit/test_agent2.py","test_lit_score_five_perfect",
     "5 HNSW abstracts with similarity=0.95 produce LitScore > 0.9 (expected ~0.965).",
     "Happy Path","Unit","unit","Verifies HNSW relevance weighting"),

    ("A2-03","Agent 2 – Literature Retriever","unit/test_agent2.py","test_lit_score_one_lower_than_five",
     "1 abstract scores lower than 5 abstracts at identical similarity — volume component matters.",
     "Happy Path","Unit","unit",""),

    ("A2-04","Agent 2 – Literature Retriever","unit/test_agent2.py","test_rrf_multi_query_wins",
     "Paper appearing in 2 query result sets ranks above paper appearing in only 1.",
     "Happy Path","Unit","unit","Core RRF logic"),

    ("A2-05","Agent 2 – Literature Retriever","unit/test_agent2.py","test_rrf_no_duplicates",
     "Same PMID from two result sets appears exactly once in fused output.",
     "Happy Path","Unit","unit","De-duplication"),

    ("A2-06","Agent 2 – Literature Retriever","unit/test_agent2.py","test_rrf_keeps_best_similarity",
     "When same paper appears in two result sets, the entry with the lowest distance is kept.",
     "Happy Path","Unit","unit",""),

    ("A2-07","Agent 2 – Literature Retriever","unit/test_agent2.py","test_rrf_hnsw_and_bm25_same_paper_scores_highest",
     "Paper found by BOTH HNSW and BM25 ranks above papers found by only one retriever.",
     "Happy Path","Unit","unit","Hybrid retrieval"),

    ("A2-08","Agent 2 – Literature Retriever","integration/test_agent2.py","test_chromadb_loaded",
     "ChromaDB must contain >= 1,800 abstracts before any live retrieval test proceeds.",
     "Happy Path","Integration","live","Precondition check"),

    ("A2-09","Agent 2 – Literature Retriever","integration/test_agent2.py","test_dupilumab_conjunctivitis",
     "Live hybrid retrieval for dupilumab x conjunctivitis returns >= 3 abstracts above the 0.60 cosine threshold.",
     "Happy Path","Integration","live","Golden signal"),

    ("A2-10","Agent 2 – Literature Retriever","integration/test_agent2.py","test_gabapentin_respiratory",
     "Live hybrid retrieval for gabapentin x respiratory arrest returns valid results.",
     "Happy Path","Integration","live",""),

    ("A2-11","Agent 2 – Literature Retriever","integration/test_agent2.py","test_bm25_finds_different_papers_than_hnsw",
     "BM25 and HNSW return at least some non-overlapping papers, confirming complementary retrieval value.",
     "Happy Path","Integration","live","Hybrid diversity"),

    ("A2-12","Agent 2 – Literature Retriever","integration/test_agent2.py","test_empty_queries_graceful",
     "Agent 2 handles empty query list (Agent 1 failure scenario) without crashing — returns empty list.",
     "Sad Path – Upstream Failure","Integration","live","Graceful degradation"),

    # ── Agent 3 ──────────────────────────────────────────────────────────────
    ("A3-01","Agent 3 – Assessor","unit/test_agent3.py","test_priority_tier_p1",
     "stat_score >= 0.7 AND lit_score >= 0.5 produces priority = P1.",
     "Happy Path","Unit","unit",""),

    ("A3-02","Agent 3 – Assessor","unit/test_agent3.py","test_priority_tier_p2",
     "stat_score >= 0.7 AND lit_score < 0.5 produces priority = P2.",
     "Happy Path","Unit","unit",""),

    ("A3-03","Agent 3 – Assessor","unit/test_agent3.py","test_priority_tier_p3",
     "stat_score < 0.7 AND lit_score >= 0.5 produces priority = P3.",
     "Happy Path","Unit","unit",""),

    ("A3-04","Agent 3 – Assessor","unit/test_agent3.py","test_priority_tier_p4",
     "Both scores below threshold produces priority = P4.",
     "Happy Path","Unit","unit",""),

    ("A3-05","Agent 3 – Assessor","unit/test_agent3.py","test_priority_boundary_exact_thresholds",
     "Values exactly at boundary (stat=0.7, lit=0.5) are assigned to the higher tier (P1).",
     "Edge Case / Boundary","Unit","unit","Boundary: inclusive threshold"),

    ("A3-06","Agent 3 – Assessor","unit/test_agent3.py","test_stat_score_fallback_formula",
     "Agent 3 computes stat_score locally when Agent 1 did not supply it; PRR=4.0, 50 cases => ~0.70.",
     "Happy Path","Unit","unit","Fallback formula"),

    ("A3-07","Agent 3 – Assessor","unit/test_agent3.py","test_stat_score_death_increases_severity",
     "Adding a death flag raises the severity component of stat_score from 0.0 to 1.0.",
     "Happy Path","Unit","unit",""),

    ("A3-08","Agent 3 – Assessor","unit/test_agent3.py","test_stat_score_lt_higher_than_hosp",
     "Life-threatening events (weight 0.75) outrank hospitalisation (weight 0.50) in severity score.",
     "Happy Path","Unit","unit",""),

    ("A3-09","Agent 3 – Assessor","unit/test_agent3.py","test_pydantic_rejects_missing_fields",
     "SafetyBriefOutput rejects output with missing required fields — raises ValidationError.",
     "Sad Path – Invalid Input","Unit","unit","Schema validation"),

    ("A3-10","Agent 3 – Assessor","unit/test_agent3.py","test_pydantic_rejects_bad_recommended_action",
     "recommended_action must be one of MONITOR / LABEL_UPDATE / RESTRICT / WITHDRAW — other values rejected.",
     "Sad Path – Invalid Input","Unit","unit","Schema validation"),

    ("A3-11","Agent 3 – Assessor","unit/test_agent3.py","test_pydantic_rejects_stat_score_out_of_range",
     "stat_score outside [0.0, 1.0] is rejected by Pydantic validation.",
     "Sad Path – Invalid Input","Unit","unit","Range validation"),

    ("A3-12","Agent 3 – Assessor","unit/test_agent3.py","test_pydantic_accepts_valid_brief",
     "Correctly formed SafetyBriefOutput passes all Pydantic validation rules.",
     "Happy Path","Unit","unit",""),

    ("A3-13","Agent 3 – Assessor","unit/test_agent3.py","test_citation_guard_removes_fabricated_pmids",
     "PMIDs in the brief but absent from the retrieved abstract set are stripped before writing to Snowflake.",
     "Happy Path","Unit","unit","Hallucination guard"),

    ("A3-14","Agent 3 – Assessor","unit/test_agent3.py","test_citation_guard_allows_all_when_all_valid",
     "No PMIDs removed when every cited PMID is contained in the retrieved abstract set.",
     "Happy Path","Unit","unit",""),

    ("A3-15","Agent 3 – Assessor","unit/test_agent3.py","test_citation_guard_empty_retrieved",
     "If Agent 2 returned zero abstracts, all PMIDs cited by GPT-4o are treated as fabricated and stripped.",
     "Edge Case","Unit","unit","Empty upstream"),

    ("A3-16","Agent 3 – Assessor","unit/test_agent3.py","test_agent3_uses_state_stat_score_when_present",
     "If stat_score is already in state (from Agent 1), Agent 3 uses it verbatim without re-computing.",
     "Happy Path","Unit","unit",""),

    ("A3-17","Agent 3 – Assessor","unit/test_agent3.py","test_agent3_node_returns_required_state_keys",
     "agent3_node returns a dict with at minimum 'priority' and 'brief' keys (GPT-4o and Snowflake mocked).",
     "Happy Path","Unit","unit","Mocked IO"),

    ("A3-18","Agent 3 – Assessor","unit/test_agent3.py","test_normalize_action_maps_prose_variants",
     "_normalize_action maps GPT-4o prose variants (e.g. 'label update recommended') to valid literals.",
     "Happy Path","Unit","unit","Normalisation"),

    ("A3-19","Agent 3 – Assessor","unit/test_agent3.py","test_agent3_node_sets_gen_error_on_double_failure",
     "When both GPT-4o retry attempts return malformed JSON, agent3_node sets generation_error=True and returns brief=None.",
     "Sad Path – GPT-4o Failure","Unit","unit","Double failure"),

    ("A3-20","Agent 3 – Assessor","integration/test_agent3.py","test_agent3_full_run_bupropion",
     "End-to-end Agent 3 with real GPT-4o and Snowflake: bupropion x seizure. Verifies priority, brief written, PMIDs valid, row persisted.",
     "Happy Path","Integration","integration","Live GPT-4o + Snowflake"),

    # ── Hallucination Check ───────────────────────────────────────────────────
    ("HC-01","Hallucination Detection","unit/test_hallucination_check.py","test_numerical_accuracy_clean",
     "Brief with correct PRR, case count, and death count produces hallucination_rate = 0.0.",
     "Happy Path","Unit","none",""),

    ("HC-02","Hallucination Detection","unit/test_hallucination_check.py","test_numerical_accuracy_hallucinated_prr",
     "Brief claims PRR = 10.5 but actual PRR = 3.5 — fabricated value detected, rate > 0.",
     "Sad Path – Hallucination","Unit","none","PRR fabrication"),

    ("HC-03","Hallucination Detection","unit/test_hallucination_check.py","test_numerical_accuracy_false_death_claim",
     "Brief says 'no deaths reported' but data shows death_count = 5 — false claim detected.",
     "Sad Path – Hallucination","Unit","none","Death count lie"),

    ("HC-04","Hallucination Detection","unit/test_hallucination_check.py","test_priority_action_consistency_valid",
     "P1 signal with LABEL_UPDATE and serious outcomes — consistent, hallucination_rate = 0.0.",
     "Happy Path","Unit","none",""),

    ("HC-05","Hallucination Detection","unit/test_hallucination_check.py","test_priority_action_consistency_restrict_with_low_prr_deaths",
     "RESTRICT with deaths and PRR = 2.98 (above the 2.0 threshold) — valid recommendation.",
     "Happy Path","Unit","none","Boundary: PRR just above 2"),

    ("HC-06","Hallucination Detection","unit/test_hallucination_check.py","test_priority_action_consistency_invalid_withdraw",
     "WITHDRAW recommended for P4 with 0 deaths and PRR = 2.1 — three errors, rate = 1.0.",
     "Sad Path – Hallucination","Unit","none","Unjustified WITHDRAW"),

    ("HC-07","Hallucination Detection","unit/test_hallucination_check.py","test_priority_action_consistency_monitor_mild_reaction",
     "P1 signal with MONITOR for mild reaction (0 deaths, 0 LT, 0 hosp) — valid action, rate = 0.0.",
     "Happy Path","Unit","none","Mild AE exception"),

    ("HC-08","Hallucination Detection","unit/test_hallucination_check.py","test_citation_grounding_no_abstracts",
     "No abstracts available — grounding check skipped, hallucination_rate = 0.0 (data issue, not hallucination).",
     "Edge Case","Unit","none",""),

    ("HC-09","Hallucination Detection","unit/test_hallucination_check.py","test_citation_grounding_with_valid_citations",
     "Brief cites a PMID with high semantic similarity to retrieved abstract — rate < 0.5.",
     "Happy Path","Unit","none","Citation grounding"),

    ("HC-10","Hallucination Detection","unit/test_hallucination_check.py","test_composite_validation_clean_brief",
     "Full validate_brief() on clean dupilumab x conjunctivitis brief — pass=True, composite score < 0.20.",
     "Happy Path","Unit","none","Composite check"),

    ("HC-11","Hallucination Detection","unit/test_hallucination_check.py","test_composite_validation_problematic_brief",
     "Full validate_brief() on brief with fabricated PRR=15, 500 cases, WITHDRAW at P4 — pass=False, score > 0.20.",
     "Sad Path – Hallucination","Unit","none","Multiple hallucinations"),

    # ── HITL ──────────────────────────────────────────────────────────────────
    ("HT-01","HITL Queue","unit/test_hitl.py","test_health_returns_ok",
     "GET /health returns HTTP 200 with status='ok' (Snowflake mocked).",
     "Happy Path","Unit","unit",""),

    ("HT-02","HITL Queue","unit/test_hitl.py","test_health_is_fast",
     "Health check responds in < 15 seconds (accommodates Snowflake cold-start latency).",
     "Happy Path","Unit","unit","Latency SLA"),

    ("HT-03","HITL Queue","unit/test_hitl.py","test_post_decision_rejects_invalid_decision",
     "POST /hitl/decisions with decision='MAYBE' returns 422 Unprocessable Entity.",
     "Sad Path – Invalid Input","Unit","unit","Input validation"),

    ("HT-04","HITL Queue","unit/test_hitl.py","test_post_decision_rejects_missing_drug_key",
     "POST /hitl/decisions without drug_key field returns 422.",
     "Sad Path – Missing Field","Unit","unit",""),

    ("HT-05","HITL Queue","unit/test_hitl.py","test_post_decision_rejects_missing_pt",
     "POST /hitl/decisions without pt field returns 422.",
     "Sad Path – Missing Field","Unit","unit",""),

    ("HT-06","HITL Queue","unit/test_hitl.py","test_post_decision_accepts_lowercase",
     "Decision value 'approve' (lowercase) is accepted and normalised to 'APPROVE'.",
     "Happy Path","Unit","unit","Case-insensitive"),

    ("HT-07","HITL Queue","unit/test_hitl.py","test_post_decision_reviewer_note_is_optional",
     "Omitting reviewer_note from the request body does not cause a validation error.",
     "Happy Path","Unit","unit","Optional field"),

    ("HT-08","HITL Queue","unit/test_hitl.py","test_hitl_decision_model_all_valid_decisions",
     "All three valid values APPROVE / REJECT / ESCALATE pass Pydantic validation.",
     "Happy Path","Unit","unit","All valid literals"),

    ("HT-09","HITL Queue","unit/test_hitl.py","test_hitl_decision_model_brief_id_is_optional",
     "brief_id is Optional — omitting it produces no validation error.",
     "Happy Path","Unit","unit","Optional field"),

    ("HT-10","HITL Queue","unit/test_hitl.py","test_get_queue_returns_list",
     "GET /hitl/queue always returns a JSON list, never a dict or null.",
     "Happy Path","Unit","unit",""),

    ("HT-11","HITL Queue","unit/test_hitl.py","test_get_queue_returns_correct_shape",
     "Every item in the queue list contains the required fields including brief_id.",
     "Happy Path","Unit","unit","Schema shape"),

    ("HT-12","HITL Queue","unit/test_hitl.py","test_get_decisions_returns_list",
     "GET /hitl/decisions always returns a JSON list.",
     "Happy Path","Unit","unit",""),

    ("HT-13","HITL Queue","integration/test_hitl.py","test_health_endpoint_live",
     "Live health check connects to real Snowflake and returns snowflake_version as connectivity proof.",
     "Happy Path","Integration","integration","Live Snowflake"),

    ("HT-14","HITL Queue","integration/test_hitl.py","test_get_queue_live",
     "Live GET /hitl/queue returns a list of pending signals from real Snowflake.",
     "Happy Path","Integration","integration",""),

    ("HT-15","HITL Queue","integration/test_hitl.py","test_get_decisions_live",
     "Live GET /hitl/decisions returns full audit log from real Snowflake.",
     "Happy Path","Integration","integration",""),

    ("HT-16","HITL Queue","integration/test_hitl.py","test_post_decision_live",
     "Live POST /hitl/decisions writes a real row to hitl_decisions and verifies with follow-up query.",
     "Happy Path","Integration","integration","Writes to Snowflake"),

    # ── Evaluation ────────────────────────────────────────────────────────────
    ("EV-01","Evaluation – Golden Signals","unit/test_evaluation.py","test_exactly_10_golden_signals",
     "GOLDEN_SIGNALS constant contains exactly 10 entries.",
     "Happy Path","Unit","none",""),

    ("EV-02","Evaluation – Golden Signals","unit/test_evaluation.py","test_all_required_fields_present",
     "Every golden signal record has all required fields: drug_key, pt, fda_comm_date, fda_comm_label.",
     "Happy Path","Unit","none","Schema completeness"),

    ("EV-03","Evaluation – Golden Signals","unit/test_evaluation.py","test_all_10_golden_drugs_present",
     "All 10 expected drug keys are present in the golden signal set.",
     "Happy Path","Unit","none",""),

    ("EV-04","Evaluation – Golden Signals","unit/test_evaluation.py","test_fda_comm_dates_are_date_objects",
     "fda_comm_date values are Python date objects, not strings — type guard.",
     "Sad Path – Type Error","Unit","none","Type guard"),

    ("EV-05","Evaluation – Golden Signals","unit/test_evaluation.py","test_fda_comm_dates_in_2023_or_2024",
     "All FDA communication dates fall in 2023 or 2024 (FAERS data year range).",
     "Happy Path","Unit","none",""),

    ("EV-06","Evaluation – Golden Signals","unit/test_evaluation.py","test_no_duplicate_drug_pt_pairs",
     "No two golden signals share the same (drug_key, pt) pair.",
     "Happy Path","Unit","none","Uniqueness"),

    ("EV-07","Evaluation – Golden Signals","unit/test_evaluation.py","test_pt_values_are_lowercase",
     "All pt values in the golden set are lowercase (must match FAERS normalisation).",
     "Happy Path","Unit","none","Normalisation"),

    ("EV-08","Evaluation – Golden Signals","unit/test_evaluation.py","test_drug_keys_are_lowercase",
     "All drug_key values in the golden set are lowercase.",
     "Happy Path","Unit","none","Normalisation"),

    ("EV-09","Evaluation – Golden Signals","unit/test_evaluation.py","test_fda_comm_labels_non_empty",
     "fda_comm_label is a non-empty string for every golden signal.",
     "Happy Path","Unit","none",""),

    ("EV-10","Evaluation – Golden Signals","unit/test_evaluation.py","test_positive_lead_time",
     "Lead-time computation returns positive days when detection precedes FDA communication.",
     "Happy Path","Unit","none",""),

    ("EV-11","Evaluation – Golden Signals","unit/test_evaluation.py","test_negative_lead_time",
     "Lead-time computation returns negative days when detection follows FDA communication.",
     "Sad Path – Late Detection","Unit","none",""),

    ("EV-12","Evaluation – Golden Signals","unit/test_evaluation.py","test_zero_lead_time",
     "Lead-time = 0 when detection date equals FDA communication date.",
     "Edge Case / Boundary","Unit","none","Same-day boundary"),

    ("EV-13","Evaluation – Golden Signals","unit/test_evaluation.py","test_metformin_lead_time",
     "Metformin POC lead time is approximately 13 days (per project proposal).",
     "Happy Path","Unit","none",""),

    ("EV-14","Evaluation – Golden Signals","unit/test_evaluation.py","test_dupilumab_lead_time",
     "Dupilumab POC lead time is approximately 291 days (per project proposal).",
     "Happy Path","Unit","none",""),

    ("EV-15","Evaluation – Golden Signals","unit/test_evaluation.py","test_precision_all_flagged",
     "Precision = 1.0 when all 10 golden signals are flagged.",
     "Happy Path","Unit","none",""),

    ("EV-16","Evaluation – Golden Signals","unit/test_evaluation.py","test_precision_six_flagged",
     "Precision = 0.6 when only 6 of 10 golden signals are flagged.",
     "Happy Path","Unit","none","Partial detection"),

    ("EV-17","Evaluation – Golden Signals","unit/test_evaluation.py","test_precision_zero_flagged",
     "Precision = 0.0 when none of the golden signals are flagged.",
     "Sad Path – Zero Detection","Unit","none","Worst-case"),

    # ── Redis ──────────────────────────────────────────────────────────────────
    ("RD-01","Redis Cache","unit/test_redis.py","test_signal_cache_key_no_priority",
     "Signal cache key without priority filter contains 'all' in the key string.",
     "Happy Path","Unit","unit",""),

    ("RD-02","Redis Cache","unit/test_redis.py","test_signal_cache_key_with_priority",
     "Signal cache key with priority='P1' contains 'P1' in the key string.",
     "Happy Path","Unit","unit",""),

    ("RD-03","Redis Cache","unit/test_redis.py","test_signal_cache_key_different_priorities_are_different",
     "P1 and P2 filters produce different cache keys (no collision).",
     "Happy Path","Unit","unit","Collision check"),

    ("RD-04","Redis Cache","unit/test_redis.py","test_signal_cache_key_different_limits_are_different",
     "Different row limits produce different cache keys.",
     "Happy Path","Unit","unit","Collision check"),

    ("RD-05","Redis Cache","unit/test_redis.py","test_brief_cache_key_format",
     "Brief cache key contains both drug_key and pt as expected.",
     "Happy Path","Unit","unit",""),

    ("RD-06","Redis Cache","unit/test_redis.py","test_brief_cache_key_spaces_handled",
     "Spaces in pt value do not break the cache key (no KeyError or encoding error).",
     "Edge Case","Unit","unit","Space handling"),

    ("RD-07","Redis Cache","unit/test_redis.py","test_brief_cache_key_different_signals_are_different",
     "Two different drug/pt pairs produce different cache keys.",
     "Happy Path","Unit","unit","Collision check"),

    ("RD-08","Redis Cache","integration/test_redis.py","test_redis_connection",
     "Redis is reachable on localhost:6379 — ping returns True.",
     "Happy Path","Integration","live","Precondition"),

    ("RD-09","Redis Cache","integration/test_redis.py","test_cache_set_and_get",
     "Value stored with cache_set() is retrievable with cache_get() and all fields match.",
     "Happy Path","Integration","live",""),

    ("RD-10","Redis Cache","integration/test_redis.py","test_cache_get_missing_key_returns_none",
     "cache_get() for a non-existent key returns None — no exception raised.",
     "Sad Path – Missing Key","Integration","live","Graceful miss"),

    ("RD-11","Redis Cache","integration/test_redis.py","test_cache_delete",
     "Key is no longer retrievable after cache_delete() is called.",
     "Happy Path","Integration","live",""),

    ("RD-12","Redis Cache","integration/test_redis.py","test_cache_ttl_respected",
     "Value stored with TTL=2s expires and is gone after 3 seconds.",
     "Happy Path","Integration","live","TTL expiry"),

    ("RD-13","Redis Cache","integration/test_redis.py","test_signal_cache_full_flow",
     "Full signal cache lifecycle: miss -> store -> hit -> invalidate_signals() -> miss.",
     "Happy Path","Integration","live","Full flow"),

    ("RD-14","Redis Cache","integration/test_redis.py","test_brief_cache_full_flow",
     "Full brief cache lifecycle: miss -> store -> hit -> invalidate_brief() -> miss.",
     "Happy Path","Integration","live","Full flow"),

    ("RD-15","Redis Cache","integration/test_redis.py","test_queue_depth_set_and_get",
     "set_queue_depth(42) then get_queue_depth() returns 42.",
     "Happy Path","Integration","live",""),

    ("RD-16","Redis Cache","integration/test_redis.py","test_queue_depth_updates",
     "Queue depth update from 100 to 99 is correctly stored (old value overwritten).",
     "Happy Path","Integration","live",""),

    ("RD-17","Redis Cache","integration/test_redis.py","test_invalidate_signals_clears_all_priorities",
     "invalidate_signals() clears ALL priority-filter keys (P1, P2, and no-filter) in a single call.",
     "Happy Path","Integration","live","Bulk invalidation"),

    # ── On-Demand Pipeline ────────────────────────────────────────────────────
    ("OD-01","On-Demand Pipeline","integration/test_on_demand.py","test_known_signal_completes",
     "run_single_signal('dupilumab', 'conjunctivitis') completes without exception or generation_error.",
     "Happy Path","Integration","integration","Golden signal"),

    ("OD-02","On-Demand Pipeline","integration/test_on_demand.py","test_priority_tier_valid",
     "dupilumab x conjunctivitis returns priority = P1.",
     "Happy Path","Integration","integration","Known P1"),

    ("OD-03","On-Demand Pipeline","integration/test_on_demand.py","test_brief_generated",
     "SafetyBrief is not None and contains brief_text, key_findings, pmids_cited, recommended_action.",
     "Happy Path","Integration","integration",""),

    ("OD-04","On-Demand Pipeline","integration/test_on_demand.py","test_brief_content_valid",
     "brief_text mentions 'dupilumab' and 'conjunctivitis'; key_findings non-empty; recommended_action is a valid literal.",
     "Happy Path","Integration","integration","Clinical correctness"),

    ("OD-05","On-Demand Pipeline","integration/test_on_demand.py","test_scores_in_range",
     "stat_score and lit_score are both in [0.0, 1.0].",
     "Happy Path","Integration","integration","Score bounds"),

    ("OD-06","On-Demand Pipeline","integration/test_on_demand.py","test_snowflake_updated",
     "After the run, Snowflake safety_briefs has a row for dupilumab x conjunctivitis with valid priority and generation_error=False.",
     "Happy Path","Integration","integration","Persistence"),

    ("OD-07","On-Demand Pipeline","integration/test_on_demand.py","test_nonexistent_signal_raises",
     "run_single_signal('aspirin', 'headache') raises ValueError containing 'not found' (signal absent from signals_flagged).",
     "Sad Path – Invalid Signal","Integration","integration","Signal not in DB"),

    ("OD-08","On-Demand Pipeline","integration/test_on_demand.py","test_nongolden_signal_completes",
     "run_single_signal('tenofovir disoproxil', 'renal injury') completes without error — triggers on-demand PubMed fetch.",
     "Happy Path","Integration","integration","Non-golden / on-demand fetch"),

    ("OD-09","On-Demand Pipeline","integration/test_on_demand.py","test_nongolden_brief_generated",
     "Brief for tenofovir disoproxil x renal injury mentions 'tenofovir' and has a valid recommended_action.",
     "Happy Path","Integration","integration",""),

    # ── Snowflake Connectivity ─────────────────────────────────────────────────
    ("SF-01","Snowflake Connector","integration/test_snowflake_connector.py","test_python_connector",
     "snowflake-connector-python connects to Snowflake and executes SELECT CURRENT_VERSION() successfully.",
     "Happy Path","Integration","integration","Connectivity"),

    ("SF-02","Snowflake Connector","integration/test_snowflake_connector.py","test_spark_write",
     "Spark writes a 3-row test DataFrame to Snowflake via JDBC without error.",
     "Happy Path","Integration","integration","Write path"),

    ("SF-03","Snowflake Connector","integration/test_snowflake_connector.py","test_spark_read",
     "Spark reads back the table written in test_spark_write and row count matches (round-trip verification).",
     "Happy Path","Integration","integration","Read path"),
]

# ── Path-type -> row fill colour ─────────────────────────────────────────────
PATH_FILLS = {
    "Happy Path"                  : fill("C8E6C9"),
    "Sad Path – Invalid Input"    : fill("FFCDD2"),
    "Sad Path – Hallucination"    : fill("FFCDD2"),
    "Sad Path – GPT-4o Failure"   : fill("FFCDD2"),
    "Sad Path – Late Detection"   : fill("FFCDD2"),
    "Sad Path – Zero Detection"   : fill("FFCDD2"),
    "Sad Path – Missing Key"      : fill("FFCDD2"),
    "Sad Path – Invalid Signal"   : fill("FFCDD2"),
    "Sad Path – Upstream Failure" : fill("FFCDD2"),
    "Sad Path – Missing Field"    : fill("FFCDD2"),
    "Sad Path – Type Error"       : fill("FFCDD2"),
    "Edge Case"                   : fill("FFF9C4"),
    "Edge Case / Boundary"        : fill("FFF9C4"),
}

# ── Build workbook ────────────────────────────────────────────────────────────
wb = Workbook()

# ============================================================
# Sheet 1: Master Catalogue
# ============================================================
ws = wb.active
ws.title = "Master Catalogue"

# Title banner
ws.merge_cells("A1:I1")
ws["A1"].value     = "MedSignal Platform  —  Test Scenario Catalogue"
ws["A1"].font      = Font(bold=True, size=14, color="FFFFFF")
ws["A1"].fill      = fill(GREEN_DARK)
ws["A1"].alignment = center()
ws.row_dimensions[1].height = 32

# Sub-title
ws.merge_cells("A2:I2")
ws["A2"].value     = "Happy Paths  |  Sad Paths  |  Edge Cases    Unit + Integration"
ws["A2"].font      = Font(italic=True, size=10, color="1B5E20")
ws["A2"].fill      = fill("F1F8E9")
ws["A2"].alignment = center()
ws.row_dimensions[2].height = 18

ws.append([])  # blank spacer

# Column headers
HEADERS = [
    "Test ID", "Component", "Test File", "Test Function",
    "Description", "Path Type", "Level", "Marker", "Notes"
]
ws.append(HEADERS)
for col_idx, h in enumerate(HEADERS, start=1):
    cell           = ws.cell(row=4, column=col_idx)
    cell.value     = h
    cell.font      = bold_font(10, "FFFFFF")
    cell.fill      = fill(GREEN_MID)
    cell.alignment = center(wrap=False)
    cell.border    = thin_border()
ws.row_dimensions[4].height = 22

# Column widths
COL_WIDTHS = [9, 26, 36, 44, 72, 28, 14, 14, 36]
for i, w in enumerate(COL_WIDTHS, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Data rows
for r_idx, row in enumerate(TESTS, start=5):
    path_type = row[5]
    level     = row[6]
    row_fill  = PATH_FILLS.get(path_type, fill("FFFFFF"))

    for c_idx, val in enumerate(row, start=1):
        cell           = ws.cell(row=r_idx, column=c_idx, value=val)
        cell.border    = thin_border()
        cell.fill      = row_fill
        cell.alignment = left_align(wrap=True) if c_idx in (4, 5, 9) else center(wrap=True)

        # Level column: blue for Integration
        if c_idx == 7 and level == "Integration":
            cell.fill = fill("BBDEFB")
            cell.font = Font(bold=True, size=10, color="0D47A1")

    ws.row_dimensions[r_idx].height = 40

ws.freeze_panes = "A5"

# ============================================================
# Sheet 2: Summary by Component
# ============================================================
ws2 = wb.create_sheet("Summary by Component")

ws2.merge_cells("A1:G1")
ws2["A1"].value     = "Summary  —  Test Coverage by Component"
ws2["A1"].font      = Font(bold=True, size=13, color="FFFFFF")
ws2["A1"].fill      = fill(GREEN_DARK)
ws2["A1"].alignment = center()
ws2.row_dimensions[1].height = 28

SUMM_HEADERS = ["Component","Unit Tests","Integration Tests","Happy Paths","Sad Paths","Edge Cases","Total"]
ws2.append(SUMM_HEADERS)
for c_idx, h in enumerate(SUMM_HEADERS, start=1):
    cell           = ws2.cell(row=2, column=c_idx)
    cell.font      = bold_font(10, "FFFFFF")
    cell.fill      = fill(GREEN_MID)
    cell.alignment = center()
    cell.border    = thin_border()
ws2.row_dimensions[2].height = 22

components: dict = {}
for row in TESTS:
    comp      = row[1]
    path_type = row[5]
    level     = row[6]
    if comp not in components:
        components[comp] = {"unit": 0, "integration": 0, "happy": 0, "sad": 0, "edge": 0}
    if level == "Unit":
        components[comp]["unit"] += 1
    else:
        components[comp]["integration"] += 1
    if path_type.startswith("Happy"):
        components[comp]["happy"] += 1
    elif path_type.startswith("Sad"):
        components[comp]["sad"] += 1
    else:
        components[comp]["edge"] += 1

totals_acc = {"unit": 0, "integration": 0, "happy": 0, "sad": 0, "edge": 0}
for r_idx, (comp, counts) in enumerate(components.items(), start=3):
    row_total = counts["unit"] + counts["integration"]
    row_data  = [comp, counts["unit"], counts["integration"],
                 counts["happy"], counts["sad"], counts["edge"], row_total]
    for c_idx, val in enumerate(row_data, start=1):
        cell           = ws2.cell(row=r_idx, column=c_idx, value=val)
        cell.fill      = fill("F9FBE7")
        cell.border    = thin_border()
        cell.alignment = center(wrap=True)
    ws2.row_dimensions[r_idx].height = 20
    for k in totals_acc:
        totals_acc[k] += counts[k]
    last_data_row = r_idx

tot_row = last_data_row + 1
grand_total = totals_acc["unit"] + totals_acc["integration"]
for c_idx, val in enumerate(
    ["TOTAL", totals_acc["unit"], totals_acc["integration"],
     totals_acc["happy"], totals_acc["sad"], totals_acc["edge"], grand_total],
    start=1
):
    cell           = ws2.cell(row=tot_row, column=c_idx, value=val)
    cell.font      = Font(bold=True, size=11, color="FFFFFF")
    cell.fill      = fill(GREEN_DARK)
    cell.border    = thin_border()
    cell.alignment = center()
ws2.row_dimensions[tot_row].height = 24

SUMM_WIDTHS = [34, 14, 20, 14, 12, 14, 10]
for i, w in enumerate(SUMM_WIDTHS, start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# Sheet 3: Legend
# ============================================================
ws3 = wb.create_sheet("Legend")

ws3.merge_cells("A1:D1")
ws3["A1"].value     = "Legend  —  Colour Key and Pytest Markers"
ws3["A1"].font      = Font(bold=True, size=13, color="FFFFFF")
ws3["A1"].fill      = fill(GREEN_DARK)
ws3["A1"].alignment = center()
ws3.row_dimensions[1].height = 26

legend_colours = [
    ("Happy Path",    "C8E6C9", "Expected success — valid inputs, normal flow, correct output."),
    ("Sad Path",      "FFCDD2", "Failure scenario — invalid input, error condition, hallucination, graceful degradation."),
    ("Edge Case",     "FFF9C4", "Boundary condition, empty input, zero value, or unusual-but-valid scenario."),
    ("Integration",   "BBDEFB", "Level column highlight — test requires a live external service."),
]

ws3.append(["Colour", "Hex Code", "Category", "Meaning"])
for c_idx in range(1, 5):
    cell           = ws3.cell(row=2, column=c_idx)
    cell.font      = bold_font(10, "FFFFFF")
    cell.fill      = fill(GREEN_MID)
    cell.border    = thin_border()
    cell.alignment = center()
ws3.row_dimensions[2].height = 20

for r_idx, (label, hex_c, meaning) in enumerate(legend_colours, start=3):
    ws3.cell(row=r_idx, column=1).fill      = fill(hex_c)
    ws3.cell(row=r_idx, column=1).border    = thin_border()
    ws3.cell(row=r_idx, column=2).value     = hex_c
    ws3.cell(row=r_idx, column=2).border    = thin_border()
    ws3.cell(row=r_idx, column=2).alignment = center()
    ws3.cell(row=r_idx, column=3).value     = label
    ws3.cell(row=r_idx, column=3).font      = Font(bold=True, size=10)
    ws3.cell(row=r_idx, column=3).border    = thin_border()
    ws3.cell(row=r_idx, column=3).alignment = center()
    ws3.cell(row=r_idx, column=4).value     = meaning
    ws3.cell(row=r_idx, column=4).border    = thin_border()
    ws3.cell(row=r_idx, column=4).alignment = left_align()
    ws3.row_dimensions[r_idx].height = 20

marker_start = r_idx + 2
ws3.merge_cells(f"A{marker_start}:D{marker_start}")
ws3.cell(row=marker_start, column=1).value     = "Pytest Markers"
ws3.cell(row=marker_start, column=1).font      = Font(bold=True, size=11, color="FFFFFF")
ws3.cell(row=marker_start, column=1).fill      = fill(GREEN_MID)
ws3.cell(row=marker_start, column=1).alignment = center()
ws3.row_dimensions[marker_start].height = 20

marker_data = [
    ("@pytest.mark.unit",        "Pure logic — no network calls, no DB, no API.  Run: poetry run pytest -m unit"),
    ("@pytest.mark.live",        "Requires live ChromaDB or Redis.  Run: poetry run pytest -m live"),
    ("@pytest.mark.integration", "Requires Snowflake + OpenAI credentials.  Run: poetry run pytest -m integration"),
    ("none",                     "Custom test runner or unmarked test (not collected by default -m unit filter)."),
]

for i, (marker, meaning) in enumerate(marker_data, start=1):
    rr = marker_start + i
    ws3.cell(row=rr, column=1).value     = marker
    ws3.cell(row=rr, column=1).font      = Font(name="Courier New", size=10)
    ws3.cell(row=rr, column=1).fill      = fill("ECEFF1")
    ws3.cell(row=rr, column=1).border    = thin_border()
    ws3.cell(row=rr, column=1).alignment = left_align()
    ws3.cell(row=rr, column=2).value     = meaning
    ws3.cell(row=rr, column=2).border    = thin_border()
    ws3.cell(row=rr, column=2).alignment = left_align()
    ws3.merge_cells(f"B{rr}:D{rr}")
    ws3.row_dimensions[rr].height = 20

ws3.column_dimensions["A"].width = 32
ws3.column_dimensions["B"].width = 80
ws3.column_dimensions["C"].width = 1
ws3.column_dimensions["D"].width = 1

# ── Save ──────────────────────────────────────────────────────────────────────
out_path = "docs/medsignal_test_scenarios.xlsx"
os.makedirs("docs", exist_ok=True)
wb.save(out_path)
print(f"Saved : {out_path}")
print(f"Sheets: Master Catalogue | Summary by Component | Legend")
print(f"Total test scenarios: {len(TESTS)}")
happy = sum(1 for t in TESTS if t[5].startswith("Happy"))
sad   = sum(1 for t in TESTS if t[5].startswith("Sad"))
edge  = sum(1 for t in TESTS if t[5].startswith("Edge"))
unit  = sum(1 for t in TESTS if t[6] == "Unit")
integ = sum(1 for t in TESTS if t[6] == "Integration")
print(f"  Happy paths : {happy}")
print(f"  Sad paths   : {sad}")
print(f"  Edge cases  : {edge}")
print(f"  Unit tests  : {unit}")
print(f"  Integration : {integ}")
